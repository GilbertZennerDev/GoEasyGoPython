import csv
import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook
import openpyxl

class SpreadsheetApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Simple Spreadsheet")

        # Create a frame to hold the spreadsheet
        self.spreadsheet_frame = ttk.Frame(self)
        self.spreadsheet_frame.grid(row=0, column=0, sticky=(tk.N, tk.W, tk.E, tk.S))

        # Initialize an empty list for rows
        self.rows = []
        self.create_spreadsheet()
        
        # Create a frame to hold the buttons
        self.button_frame = ttk.Frame(self)
        self.button_frame.grid(row=0, column=1, sticky=(tk.S))

        # Save button
        self.save_button = ttk.Button(self.button_frame, text="Save", command=self.save_data)
        self.save_button.grid(row=0, column=0, padx=(10, 0), pady=(10, 0))

        # Load button
        self.load_button = ttk.Button(self.button_frame, text="Load", command=self.load_data)
        self.load_button.grid(row=0, column=1, padx=(10, 0), pady=(10, 0))

    def create_spreadsheet(self):
        # Clear existing cells
        for widget in self.spreadsheet_frame.winfo_children():
            widget.destroy()

        # Create new cells
        for row_num in range(10):  # Assuming 10 rows initially
            row = []
            for col_num in range(10):  # Assuming 10 columns initially
                cell = ttk.Entry(self.spreadsheet_frame, width=5)
                cell.grid(row=row_num, column=col_num, padx=2, pady=2)
                row.append(cell)

            self.rows.append(row)

    def get_data(self):
        # Gather data from the spreadsheet
        data = []
        for row in self.rows:
            cell_values = [cell.get() for cell in row]
            if any(cell_values):  # Don't add empty rows
                data.append(cell_values)
        return data
    
    def save_data(self, filename="spreadsheet.xlsx"):
        wb = Workbook()
        ws = wb.active

        for row_num, row in enumerate(self.get_data(), start=1):
            for col_num, cell in enumerate(row, start=1):
                ws.cell(row=row_num, column=col_num).value = cell

        wb.save(filename)
        print(f"Data saved to {filename}")


    def load_data(self, filename="spreadsheet.xlsx"):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

        for row_num in range(1, ws.max_row + 1):
            new_row = []
            for col_num in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value is not None:
                    new_entry = ttk.Entry(self.spreadsheet_frame, width=5, text=str(cell_value))
                    new_row.append(new_entry)

            # Separate row indices from Entry widgets
            row_indices = [index for index in self.rows if isinstance(index, int)]
            entry_widgets = new_row

            # Increment row indices after loading new rows
            for i, index in enumerate(row_indices):
                row_indices[i] += len(entry_widgets)

            # Place the new entries in the grid
            for entry, row_index in zip(entry_widgets, row_indices):
                entry.grid(row=row_index, column=0, padx=2, pady=2)

            self.rows.extend(row_indices)  # Add row indices after new entries
            self.rows.extend(new_row)   # Add Entry widgets to the end of the list
        print(f"Data loaded from {filename}")


if __name__ == "__main__":
    app = SpreadsheetApp()
    app.mainloop()
